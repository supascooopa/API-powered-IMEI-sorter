import json
from pathlib import Path
from openpyxl import load_workbook


def imei_counter(excel_file_path):
    quantities_dict = {}
    wb = load_workbook(excel_file_path)
    sheets_names = wb.sheetnames
    for sheet in sheets_names:
        ws = wb[sheet]
        for cols in ws.iter_cols(values_only=True):
            if cols[0]:
                # getting rid of Nones in the tuple
                clean_columns = [data for data in cols[1:] if isinstance(data, int)]
                if quantities_dict.get(cols[0]):
                    quantities_dict[cols[0]] += len(clean_columns)
                else:
                    quantities_dict[cols[0]] = len(clean_columns)
    return quantities_dict

def text_writer(my_dict):
    with open("firsttry.txt", "a+") as file:
        for key in my_dict:
            file.write("Description: " + key + " " + "Amount: " + str(my_dict[key]))
            file.write("\n")

excel_files = Path.cwd().joinpath("EXCEL").glob("*.xlsx")
for files in excel_files:
    text_writer(imei_counter(excel_file_path=files))

