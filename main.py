import time

from openpyxl import Workbook, load_workbook
from datetime import datetime
from api import imei_checker
from file_manager_v101 import get_file_name


# --- Loading workbook --- #
workbook_file = get_file_name(file_extension=".xlsx")
wb = load_workbook(workbook_file)
print("currently working on " + workbook_file)
sheet_names_lst = wb.sheetnames


# --- Creating new workbook ---#
new_wb = Workbook()
new_ws = new_wb.active
new_ws.append(["IMEI 1", "IMEI 2", "BRAND", "MODEL"])
row_number = 2

# going through each page
for sheet_names in sheet_names_lst:
    ws = wb[sheet_names]
    # --- Iterating over all the columns --- #
    for cols in ws.iter_cols(values_only=True):
        if cols[0]:
            # Assigning the column header that contains the description of the phone
            column_headers = cols[0].strip()
            # Cleaning up the column tuple from None types
            clean_cols = [c_cols for c_cols in cols if c_cols is not None]
            # Usually phone descriptions end with 1 or 2 so here we filter for phones essentially
            if column_headers is not None and column_headers.endswith("1"):
                # Assigning the ONLY IMEI to be put through the IMEI info API
                first_imei = cols[1]
                imei_data = imei_checker(first_imei)
                phone_model = imei_data[1].upper()
                try:
                    phone_brand = imei_data[0].upper()
                except AttributeError:
                    phone_brand = "N/A"

                for cells in clean_cols:
                    # Checking to see if the cell only contains numbers
                    if isinstance(cells, int):
                        # Appending to the new ws and leaving the second IMEI slot empty
                        new_ws.append([cells, " ", phone_model, phone_brand])
            # filtering for the second IMEI
            elif column_headers is not None and column_headers.endswith("2"):
                for cells in cols:
                    if isinstance(cells, int):
                        # Assigning the second IMEI next to the first one we put into the ws above.
                        new_ws.cell(row=row_number, column=2, value=cells)
                        # Need to keep track of number of rows the s
                        row_number += 1


# --- Saving new Excel file -- #
todays_date = datetime.now().strftime("%d-%m-%Y")
hour_and_minute = datetime.now().strftime("%H%m%S")
try:
    new_wb.save(todays_date + ".xlsx")
except PermissionError:

    new_wb.save(todays_date+hour_and_minute+".xlsx")




